# Third‑party
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from treebeard.mp_tree import MP_Node
from django.apps import apps
from django.db import models
from import_export.utils.mp_node_helpers import MP_NODE_AUTO_FIELDS


class ImportTemplateBuilder:
    def __init__(self, app_label):
        self.app_label = app_label
        self.app_config = apps.get_app_config(app_label)
        self.workbook = openpyxl.Workbook()
        self.workbook.remove(self.workbook.active)
        # Add app label to workbook as a named value
        self.app_named_value = DefinedName(name="_app", attr_text=f'"{self.app_label}"')
        self.workbook.defined_names.add(self.app_named_value)
        self.model_fields_map = {}
        self.fk_target_fields = set()
        self.choice_fields = []
        self.table_refs = {}
        self._collect_exportable_fields()  # Cache model fields up front

    def build_workbook(self):
        # Step 1: Create worksheet with table for each model
        for model in self.app_config.get_models():
            if model._meta.managed and not model._meta.abstract:
                self.create_model_worksheet(model)

        # Step 2: Resolve foreign keys
        self.resolve_foreign_keys()

        # Step 3: Add named ranges to all fields used as foreign keys in their tables
        self.add_named_ranges_for_foreign_keys()

        # Step 4: Add validations to foreign key fields
        self.add_foreign_key_validations()

        # step 5: Add boolean field validations
        self.add_boolean_field_validations()

        # Step 6: Add choices worksheet, choice tables and validations if required
        has_choices = any(
            'choices_type' in field_info
            for fields in self.model_fields_map.values()
            for field_info in fields
        )
        if has_choices:
            self.add_choices_sheet()
            self._add_choice_field_validations()

        # Step 7: Add parent field validation for MP_Node models
        self._add_parent_field_validations_for_mp_node_models()

        # Step 8:
        # self.format_special_fields()

        return self.workbook

    def get_exportable_fields(self, model):
        export_fields = []

        for field in model._meta.get_fields():
            if field.name in MP_NODE_AUTO_FIELDS:
                continue
            if isinstance(field, models.Field) and not field.auto_created:
                field_meta = {
                    'field_name': field.name,
                    'header': field.name,
                    'nullable': field.null
                }
                if isinstance(field, models.ForeignKey):
                    related_model = field.related_model
                    unique_constraints = [
                        c for c in related_model._meta.constraints
                        if isinstance(c, models.UniqueConstraint)
                    ]
                    if unique_constraints:
                        unique_fields = unique_constraints[0].fields
                        for unique_field in unique_fields:
                            export_fields.append({
                                'field_name': f"{field.name} {unique_field}",
                                'header': f"{field.name}\n{unique_field}",
                                'related_model': related_model,
                                'resolved_field': unique_field,
                                'nullable': field.null
                            })
                    else:
                        export_fields.append({
                            **field_meta,
                            'related_model': related_model
                        })
                elif field.choices:
                    export_fields.append({
                        **field_meta,
                        'choices_type': list(field.choices)
                    })
                elif isinstance(field, models.BooleanField):
                    export_fields.append({
                        **field_meta,
                        'boolean_field': True
                    })
                else:
                    export_fields.append(field_meta)

        if issubclass(model, MP_Node):
            export_fields.append({
                'field_name': 'parent',
                'header': 'parent',
                'nullable': True,  # Optional to allow root node creation
                'mp_node_parent': True  # Flag for downstream validation logic
            })

        return export_fields

    def create_model_worksheet(self, model):
        model_name = model.__name__
        ws = self.workbook.create_sheet(title=model_name)
        ws.column_dimensions['A'].width = 2
        ws['A1'].value = model_name
        ws['A1'].font = Font(bold=True)

        exportable_fields = self.get_exportable_fields(model)
        self.model_fields_map[model_name] = exportable_fields

        start_col, start_row = 2, 3
        for col_index, field_info in enumerate(exportable_fields):
            col_letter = get_column_letter(start_col + col_index)
            cell = ws[f"{col_letter}{start_row}"]
            cell.value = field_info['header']
            cell.alignment = Alignment(wrap_text=True)
            ws.column_dimensions[col_letter].width = max(len(field_info['header'].split('\n')[0]) + 2, 12)

        end_col_letter = get_column_letter(start_col + len(exportable_fields) - 1)
        table_range = f"B3:{end_col_letter}4"
        table = Table(displayName=model_name, ref=table_range)
        style = TableStyleInfo(name="TableStyleLight1", showFirstColumn=False, showLastColumn=False,
                               showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)
        self.table_refs[model_name] = table

    def resolve_foreign_keys(self):
        for model in self.app_config.get_models():
            if model._meta.managed and not model._meta.abstract:
                exportable_fields = self.model_fields_map[model.__name__]

                for field_info in exportable_fields:
                    if 'related_model' in field_info:
                        related_model = field_info['related_model']
                        resolved_field = field_info.get('resolved_field')
                        if resolved_field:
                            target_field = resolved_field
                        else:
                            related_fields = self.model_fields_map.get(related_model.__name__)
                            if not related_fields:
                                related_fields = self.get_exportable_fields(related_model)
                            target_field = related_fields[0]['field_name']
                        self.fk_target_fields.add((related_model.__name__, target_field))

                    if field_info.get('choices_type'):
                        self.choice_fields.append((model.__name__, field_info))

    def add_named_ranges_for_foreign_keys(self):
        for model_name, exportable_fields in self.model_fields_map.items():
            ws = self.workbook[model_name]
            table = self.table_refs[model_name]
            for col_index, field_info in enumerate(exportable_fields):
                if (model_name, field_info['field_name']) in self.fk_target_fields:
                    named_range = f"lst{model_name}_{field_info['field_name']}"
                    # Use field_name for safe structured reference regardless of human-readable header
                    column_name = field_info['field_name']
                    range_ref = f"'{model_name}'!{table.displayName}[{column_name}]"
                    dn = DefinedName(name=named_range, attr_text=range_ref)
                    self.workbook.defined_names.add(dn)

    def add_foreign_key_validations(self):
        for model_name, exportable_fields in self.model_fields_map.items():
            ws = self.workbook[model_name]

            for col_index, field_info in enumerate(exportable_fields):
                if 'related_model' in field_info:
                    related_model = field_info['related_model']
                    resolved_field = field_info.get('resolved_field')

                    if resolved_field:
                        target_field = resolved_field
                    else:
                        related_fields = self.model_fields_map.get(related_model.__name__)
                        if not related_fields:
                            # fallback: recalculate if the related model wasn't in same app
                            related_fields = self.get_exportable_fields(related_model)
                        target_field = related_fields[0]['field_name']

                    named_range = f"lst{related_model.__name__}_{target_field}"
                    dv_formula = f"={named_range}"

                    allow_blank = field_info.get('nullable', True)

                    dv = DataValidation(
                        type="list",
                        formula1=dv_formula,
                        allow_blank=allow_blank,
                        showDropDown=False
                    )

                    col_letter = get_column_letter(2 + col_index)
                    data_range = f"{col_letter}4:{col_letter}1048576"
                    dv.add(data_range)
                    ws.add_data_validation(dv)

    def add_boolean_field_validations(self):
        for model_name, exportable_fields in self.model_fields_map.items():
            ws = self.workbook[model_name]

            for col_index, field_info in enumerate(exportable_fields):
                if field_info.get('boolean_field'):
                    dv_formula = '"TRUE,FALSE"'

                    dv = DataValidation(
                        type="list",
                        formula1=dv_formula,
                        allow_blank=field_info.get('nullable', True),
                        showDropDown=False
                    )

                    col_letter = get_column_letter(2 + col_index)
                    data_range = f"{col_letter}4:{col_letter}1048576"
                    dv.add(data_range)
                    ws.add_data_validation(dv)

    def add_choices_sheet(self):
        ws_choices = self.workbook.create_sheet(title="Choices")
        ws_choices['A1'].value = "Choices"
        ws_choices['A1'].font = Font(bold=True)
        current_col = 2

        for model_name, field_info in self.choice_fields:
            choices_name = f"{model_name}_{field_info['field_name']}_choices"
            key_col = current_col
            label_col = current_col + 1

            ws_choices[f"{get_column_letter(key_col)}3"] = f"{field_info['field_name']}_key"
            ws_choices[f"{get_column_letter(label_col)}3"] = f"{field_info['field_name']}_label"

            row = 4
            for key, label in field_info['choices_type']:
                ws_choices[f"{get_column_letter(key_col)}{row}"] = key
                ws_choices[f"{get_column_letter(label_col)}{row}"] = label
                row += 1
            end_row = row - 1

            table_range = f"{get_column_letter(key_col)}3:{get_column_letter(label_col)}{end_row}"
            table = Table(displayName=choices_name, ref=table_range)
            style = TableStyleInfo(name="TableStyleLight1", showFirstColumn=False, showLastColumn=False,
                                   showRowStripes=True, showColumnStripes=False)
            table.tableStyleInfo = style
            ws_choices.add_table(table)
            dn = DefinedName(name=f"lst{choices_name}",
                             attr_text=f"Choices!${get_column_letter(label_col)}$4:${get_column_letter(label_col)}${end_row}")
            self.workbook.defined_names.add(dn)
            current_col += 3

        for col_idx in range(2, current_col):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for cell in ws_choices[column_letter]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws_choices.column_dimensions[column_letter].width = max_length + 2

        ws_choices.column_dimensions['A'].width = 2  # Ensure column A width

    def _add_choice_field_validations(self):
        for model_name, exportable_fields in self.model_fields_map.items():
            ws = self.workbook[model_name]
            for col_index, field_info in enumerate(exportable_fields):
                if field_info.get('choices_type'):
                    named_range = f"lst{model_name}_{field_info['field_name']}_choices"
                    dv_formula = f"={named_range}"

                    dv = DataValidation(
                        type="list",
                        formula1=dv_formula,
                        allow_blank=field_info.get('nullable', True),
                        showDropDown=False
                    )

                    col_letter = get_column_letter(2 + col_index)
                    data_range = f"{col_letter}4:{col_letter}1048576"
                    dv.add(data_range)
                    ws.add_data_validation(dv)

    def _add_parent_field_validations_for_mp_node_models(self):
        for model_name, exportable_fields in self.model_fields_map.items():
            ws = self.workbook[model_name]
            table = self.table_refs[model_name]
            try:
                model = apps.get_model(self.app_label, model_name)
                if issubclass(model, MP_Node):
                    related_fields = self.model_fields_map.get(model_name)
                    if related_fields:
                        first_natural_key = related_fields[0]['field_name']
                        named_range = f"lst{model_name}_{first_natural_key}"
                        if (model_name, first_natural_key) not in self.fk_target_fields:
                            range_ref = f"'{model_name}'!{table.displayName}[{first_natural_key}]"
                            dn = DefinedName(name=named_range, attr_text=range_ref)
                            self.workbook.defined_names.add(dn)

                        for col_index, field_info in enumerate(exportable_fields):
                            if field_info.get('mp_node_parent'):
                                dv_formula = f"={named_range}"
                                dv = DataValidation(
                                    type="list",
                                    formula1=dv_formula,
                                    allow_blank=True,
                                    showDropDown=False
                                )
                                col_letter = get_column_letter(2 + col_index)
                                data_range = f"{col_letter}4:{col_letter}1048576"
                                dv.add(data_range)
                                ws.add_data_validation(dv)

            except LookupError:
                continue

    def _collect_exportable_fields(self):
        """Populates model_fields_map for all relevant models."""
        for model in self.app_config.get_models():
            if model._meta.managed and not model._meta.abstract:
                model_name = model.__name__
                self.model_fields_map[model_name] = self.get_exportable_fields(model)
