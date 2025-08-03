import traceback
from collections import defaultdict
from django.apps import apps
from django.db import models, transaction
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.table import Table
from treebeard.mp_tree import MP_Node
from import_export.utils.model_helpers import get_cleaned_field_value, resolve_foreign_key
from import_export.utils.mp_node_helpers import create_mp_node, MP_NODE_AUTO_FIELDS


class ImportWorkbook:
    def __init__(self, full_path, app_label):
        self.full_path = full_path
        self.app_label = app_label
        self.related_model = None
        self.code_obj = None
        self.key_fields= None
        self.choice_maps = defaultdict(dict)
        self.model_fields = {}

    def import_workbook(self):
        wb = load_workbook(self.full_path, data_only=True)
        self._validate_app_label(wb)
        app_config = apps.get_app_config(self.app_label)
        app_models = app_config.get_models()

        results = {
            "successes": [],
            "failures": []
        }

        try:
            for model in app_models:
                model_name = model.__name__
                if model_name not in wb.sheetnames:
                    continue

                sheet = wb[model_name]
                table = next(
                    (tbl for tbl in sheet._tables.values() if isinstance(tbl, Table) and tbl.displayName == model_name),
                    None
                )
                if not table:
                    continue

                min_col, min_row, max_col, max_row = range_boundaries(table.ref)
                headers = [sheet.cell(row=min_row, column=col).value for col in range(min_col, max_col + 1)]

                model_fields = self._get_model_fields(model)
                for field in model_fields.values():
                    if field.choices:
                        self.choice_maps[model.__name__][field.name] = dict((label, value) for value, label in field.choices)

                    if field.is_relation and (field.many_to_one or field.one_to_one):
                        self.related_model = field.remote_field.model
                        if hasattr(self.related_model.objects, 'get_natural_key'):
                            try:
                                self.code_obj = self.related_model.objects.get_by_natural_key.__code__
                                self.key_fields = list(self.code_obj.co_varnames[:self.code_obj.argcount])[1:]
                            except Exception as e:
                                error_details = traceback.format_exc()
                                results["failures"].append(f"Could not introspect natural_key() for {self.related_model.__name__} – {e}\n{error_details}")

                            for key_field in self.key_fields:
                                self.rel_field = self.related_model._meta.get(key_field)
                                if self.rel_field.choices:
                                    self.choice_maps[self.related_model.__name__][key_field] = dict((label, value) for value, label in self.rel_field.choices)

                created_count = 0
                updated_count = 0

                with transaction.atomic():
                    data_start_row = min_row + 1
                    data_end_row = max_row - (1 if table.totalsRowCount else 0)

                    for row_idx in range(data_start_row, data_end_row + 1):
                        row_values = [sheet.cell(row=row_idx, column=col).value for col in range(min_col, max_col + 1)]
                        row_data = dict(zip(headers, row_values))
                        if not any(row_data.values()):
                            continue

                        compound_fk_data = defaultdict(dict)
                        simple_fields = {}

                        for field_name, value in row_data.items():
                            if '\n' in field_name:
                                fk_field, subfield = field_name.split('\n', 1)
                                compound_fk_data[fk_field][subfield] = value
                            else:
                                simple_fields[field_name] = value

                        data = {}
                        for field_name, value in simple_fields.items():
                            if field_name not in model_fields:
                                if issubclass(model, MP_Node) and field_name == 'parent':
                                    # Resolve parent FK instance
                                    if value is not None:
                                        parent_instance = model.objects.get_by_natural_key(value)
                                    else:
                                        parent_instance = None
                                    data[field_name] = parent_instance
                                    continue

                            field = model_fields[field_name]
                            data[field_name] = get_cleaned_field_value(field, value, self.choice_maps)

                        for fk_field, subfield_map in compound_fk_data.items():
                            if fk_field not in model_fields:
                                continue
                            field = model_fields[fk_field]
                            related_model = field.remote_field.model

                            if hasattr(related_model.objects, 'get_by_natural_key'):
                                try:
                                    code_obj = related_model.objects.get_by_natural_key.__code__
                                    key_fields = list(code_obj.co_varnames[:code_obj.co_argcount])[1:]  # skip 'self'
                                except Exception as e:
                                    error_details = traceback.format_exc()
                                    results["failures"].append(f"Could not introspect get_by_natural_key() for {related_model.__name__} – {e}\n{error_details}")
                            else:
                                raise ValueError(f"{related_model.__name__} must implement get_by_natural_key()")

                            key_values = [subfield_map.get(k) for k in key_fields]

                            if all(v is None for v in key_values):
                                if not field.null:
                                    raise ValueError(
                                        f"Field '{fk_field}' does not allow null values and no data was provided.")
                                data[fk_field] = None
                                continue
                            elif any(v is None for v in key_values):
                                raise ValueError(f"Partial values for compound FK '{fk_field}': {key_values}")

                            data[fk_field] = resolve_foreign_key(field, key_values, self.choice_maps)

                        try:
                            lookup_fields = model.objects.get_by_natural_key.__code__.co_varnames[
                                            1:model.objects.get_by_natural_key.__code__.co_argcount]
                        except Exception as e:
                            error_details = traceback.format_exc()
                            results["failures"].append(f"Could not determine natural key for model {model.__name__} – {e}\n{error_details}")

                        lookup_data = {
                            field: data.get(field)
                            for field in lookup_fields
                        }
                        if any(value is None for value in lookup_data.values()):
                            continue

                        if not issubclass(model, MP_Node):
                            obj, created = model.objects.update_or_create(
                                defaults=data,
                                **lookup_data
                            )
                            if created:
                                created_count += 1
                            else:
                                updated_count += 1
                        else:
                            instance, created = create_mp_node(model=model, data=data)
                            if created:
                                created_count += 1
                            else:
                                updated_count += 1

                results["successes"].append(f"Model name: {model_name}: {created_count} created, {updated_count} updated")

        except Exception as e:
            error_details = traceback.format_exc()
            results["failures"].append(f"Model name: {model.__name__} – {e}\n{error_details}")

        return results

    def _validate_app_label(self, wb):
        if '_app' in wb.defined_names:
            app_def = wb.defined_names.get('_app')
            if app_def:
                defined_app_label = app_def.attr_text.strip('"')

            if defined_app_label and defined_app_label != self.app_label:
                raise ValueError(f"Workbook _app name '{defined_app_label}' doesn't match provided app_label '{self.app_label}'.")

    def _get_model_fields(self, model):
        field_map = {}
        for field in model._meta.fields:
            if issubclass(model, MP_Node) and field.name in MP_NODE_AUTO_FIELDS:
                continue
            if isinstance(field, models.Field) and not field.auto_created:
                field_map[field.name] = field
        return field_map
